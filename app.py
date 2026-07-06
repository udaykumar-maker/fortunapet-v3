import os, calendar
from datetime import datetime, date, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import csv, io
from openpyxl import load_workbook
from pdf_generator import generate_pi_pdf

basedir = os.path.abspath(os.path.dirname(__file__))
os.makedirs(os.path.join(basedir, 'instance'), exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-me-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL', f"sqlite:///{os.path.join(basedir, 'instance', 'app.db')}"
).replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = "Please log in."

GST_RATE       = 0.18
DISPATCH_UNITS = ["FORTUNAPET", "UNIT 1", "DABASPETE", "MAHIMAPURA", "DADRA"]
UOM_OPTIONS    = ["Nos", "KG", "Pieces", "BAG", "TON", "BOX"]
DOC_STATUSES   = ["pending", "delivered", "lost", "price_approval_pending", "purchase_with_another_vendor"]
STATUS_LABELS  = {
    "pending": "Pending",
    "delivered": "Delivered",
    "lost": "Lost",
    "price_approval_pending": "Price Approval Pending",
    "purchase_with_another_vendor": "Purchase With Another Vendor",
}

# ── MODELS ────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    username       = db.Column(db.String(80), unique=True, nullable=False)
    full_name      = db.Column(db.String(120), nullable=False)
    password_hash  = db.Column(db.String(255), nullable=False)
    role           = db.Column(db.String(20), nullable=False, default='staff')
    active         = db.Column(db.Boolean, default=True)
    monthly_target = db.Column(db.Float, default=0)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    def set_password(self, pw):   self.password_hash = generate_password_hash(pw)
    def check_password(self, pw): return check_password_hash(self.password_hash, pw)
    def is_admin(self):           return self.role == 'admin'


class UserMonthlyTarget(db.Model):
    """Per-user, per-month targets."""
    id      = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    year    = db.Column(db.Integer, nullable=False)
    month   = db.Column(db.Integer, nullable=False)   # 1-12
    target  = db.Column(db.Float, default=0)
    user    = db.relationship('User', backref='monthly_targets')


class Customer(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    code           = db.Column(db.String(20), unique=True)
    name           = db.Column(db.String(200), nullable=False)
    contact_person = db.Column(db.String(120))
    phone          = db.Column(db.String(30))
    gst_number     = db.Column(db.String(20))
    created_by_id  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    created_by     = db.relationship('User', backref='customers')


class Item(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    name          = db.Column(db.String(200), nullable=False)
    default_uom   = db.Column(db.String(20))
    default_price = db.Column(db.Float, default=0)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    created_by    = db.relationship('User', backref='items')


class Document(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    doc_type        = db.Column(db.String(10), default='PI')
    quote_no        = db.Column(db.String(20), unique=True, nullable=False)
    doc_date        = db.Column(db.Date, default=date.today)
    dispatch_from   = db.Column(db.String(50))
    customer_id     = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    gst_applied     = db.Column(db.Boolean, default=False)
    freight_charges = db.Column(db.Float, default=0)
    base_amount     = db.Column(db.Float, default=0)
    gst_amount      = db.Column(db.Float, default=0)
    total_amount    = db.Column(db.Float, default=0)
    follow_up_date  = db.Column(db.Date)
    status          = db.Column(db.String(40), default='pending')
    lost_reason     = db.Column(db.Text)
    notes           = db.Column(db.Text)
    created_by_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at      = db.Column(db.DateTime, default=datetime.utcnow)
    # legacy columns
    item_desc  = db.Column(db.String(300))
    packaging  = db.Column(db.String(50))
    qty        = db.Column(db.Float, default=0)
    uom        = db.Column(db.String(20))
    price      = db.Column(db.Float, default=0)

    customer   = db.relationship('Customer', backref='documents')
    created_by = db.relationship('User', backref='documents')
    line_items = db.relationship('DocumentItem', backref='document',
                                 cascade='all, delete-orphan', order_by='DocumentItem.id')

    def recalc(self):
        self.base_amount  = round(sum(i.line_total for i in self.line_items), 2)
        self.gst_amount   = round(self.base_amount * GST_RATE, 2) if self.gst_applied else 0
        self.total_amount = round(self.base_amount + self.gst_amount + (self.freight_charges or 0), 2)

    @property
    def items_summary(self):
        names = [i.item_desc for i in self.line_items if i.item_desc]
        if not names: return self.item_desc or '—'
        return names[0] + (f' +{len(names)-1} more' if len(names) > 1 else '')

    @property
    def status_label(self):
        return STATUS_LABELS.get(self.status, self.status.replace('_',' ').title())


class DocumentItem(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    document_id = db.Column(db.Integer, db.ForeignKey('document.id'), nullable=False)
    item_desc   = db.Column(db.String(300), nullable=False)
    packaging   = db.Column(db.Float, default=0)   # numeric now for calculation
    qty         = db.Column(db.Float, default=0)
    uom         = db.Column(db.String(20))
    price       = db.Column(db.Float, default=0)

    @property
    def total_qty(self):
        return round((self.packaging or 0) * (self.qty or 0), 2)

    @property
    def line_total(self):
        return round((self.qty or 0) * (self.price or 0), 2)


class Counter(db.Model):
    id    = db.Column(db.Integer, primary_key=True)
    name  = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Integer, nullable=False)


def next_quote_number():
    c = Counter.query.filter_by(name='quote_no').first()
    if not c:
        c = Counter(name='quote_no', value=260001); db.session.add(c)
    else:
        c.value += 1
    db.session.commit()
    return str(c.value)


def next_customer_code():
    c = Counter.query.filter_by(name='cust_code').first()
    if not c:
        c = Counter(name='cust_code', value=100001); db.session.add(c)
    else:
        c.value += 1
    db.session.commit()
    return str(c.value)


@login_manager.user_loader
def load_user(uid): return User.query.get(int(uid))


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash("Admin access required.", "danger")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapper


def visible_docs():
    if current_user.is_admin(): return Document.query
    return Document.query.filter_by(created_by_id=current_user.id)

def visible_custs():
    if current_user.is_admin(): return Customer.query
    return Customer.query.filter_by(created_by_id=current_user.id)

def _parse_date(v):
    if not v: return None
    try: return datetime.strptime(v, '%Y-%m-%d').date()
    except: return None

def _save_line_items(doc, form):
    descs  = form.getlist('item_desc[]')
    pkgs   = form.getlist('packaging[]')
    qtys   = form.getlist('qty[]')
    uoms   = form.getlist('uom[]')
    prices = form.getlist('price[]')
    DocumentItem.query.filter_by(document_id=doc.id).delete()
    for i, desc in enumerate(descs):
        desc = desc.strip()
        if not desc: continue
        db.session.add(DocumentItem(
            document_id = doc.id,
            item_desc   = desc,
            packaging   = float(pkgs[i])   if i < len(pkgs)   and pkgs[i]   else 0,
            qty         = float(qtys[i])   if i < len(qtys)   and qtys[i]   else 0,
            uom         = uoms[i]          if i < len(uoms)                  else '',
            price       = float(prices[i]) if i < len(prices) and prices[i] else 0,
        ))

def _auto_update_lost():
    """Auto-mark PIs older than 30 days (non-delivered) as lost."""
    cutoff = date.today() - timedelta(days=30)
    old = Document.query.filter(
        Document.doc_date <= cutoff,
        Document.status.notin_(['delivered', 'lost'])
    ).all()
    for d in old:
        d.status = 'lost'
        if not d.lost_reason:
            d.lost_reason = 'Auto-closed after 30 days'
    if old: db.session.commit()

# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form.get('username','').strip()).first()
        if u and u.active and u.check_password(request.form.get('password','')):
            login_user(u); return redirect(url_for('dashboard'))
        flash("Invalid username or password.", "danger")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user(); return redirect(url_for('login'))

# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    _auto_update_lost()
    staff_filter = request.args.get('staff_id', type=int)
    all_staff    = User.query.filter_by(role='staff', active=True).order_by(User.full_name).all()

    if current_user.is_admin() and staff_filter:
        docs = Document.query.filter_by(created_by_id=staff_filter).all()
    else:
        docs = visible_docs().all()

    pending   = [d for d in docs if d.status not in ('delivered','lost')]
    delivered = [d for d in docs if d.status == 'delivered']
    lost      = [d for d in docs if d.status == 'lost']
    today     = date.today()
    overdue   = [d for d in pending if d.follow_up_date and d.follow_up_date < today]
    due_soon  = [d for d in pending if d.follow_up_date and 0 <= (d.follow_up_date-today).days <= 2]

    # Per-staff delivered totals (admin view)
    staff_totals = []
    if current_user.is_admin():
        for s in all_staff:
            sdocs = Document.query.filter_by(created_by_id=s.id).all()
            staff_totals.append({
                'name': s.full_name,
                'delivered': sum(d.total_amount or 0 for d in sdocs if d.status=='delivered'),
                'total': sum(d.total_amount or 0 for d in sdocs),
            })

    return render_template('dashboard.html',
        total_amount    = sum(d.total_amount or 0 for d in docs),
        pending_amount  = sum(d.total_amount or 0 for d in pending),
        delivered_amount= sum(d.total_amount or 0 for d in delivered),
        delivered_count = len(delivered),
        pending_count   = len(pending),
        lost_count      = len(lost),
        overdue=overdue, due_soon=due_soon,
        pending_docs=sorted(pending, key=lambda d: d.follow_up_date or date.max)[:10],
        today=today, all_staff=all_staff, staff_filter=staff_filter,
        staff_totals=staff_totals)

# ── PI / DOCUMENTS ────────────────────────────────────────────────────────────

@app.route('/pi')
@login_required
def documents():
    _auto_update_lost()
    q   = request.args.get('q','').strip()
    sf  = request.args.get('status','')
    qry = visible_docs()
    if sf: qry = qry.filter(Document.status == sf)
    docs = qry.order_by(Document.created_at.desc()).all()
    if q:
        ql   = q.lower()
        docs = [d for d in docs if
                ql in (d.customer.name.lower() if d.customer else '')
                or ql in d.quote_no.lower()
                or any(ql in (li.item_desc or '').lower() for li in d.line_items)
                or ql in (d.item_desc or '').lower()]
    custs = visible_custs().order_by(Customer.name).all()
    items = Item.query.order_by(Item.name).all()
    # build items lookup for JS autocomplete {name: {uom, price}}
    items_data = {it.name: {'uom': it.default_uom or '', 'price': it.default_price or 0} for it in items}
    return render_template('documents.html',
        docs=docs, customers=custs, items=items, items_data=items_data,
        dispatch_units=DISPATCH_UNITS, uom_options=UOM_OPTIONS,
        doc_statuses=DOC_STATUSES, status_labels=STATUS_LABELS,
        q=q, status_filter=sf,
        today=date.today(),
        today_iso=date.today().isoformat(),
        default_followup=(date.today()+timedelta(days=7)).isoformat())


@app.route('/pi/new', methods=['POST'])
@login_required
def new_document():
    cid  = request.form.get('customer_id')
    disp = request.form.get('dispatch_from')
    if not cid or not disp:
        flash("Customer and dispatch location are required.", "danger")
        return redirect(url_for('documents'))
    descs = [d.strip() for d in request.form.getlist('item_desc[]') if d.strip()]
    if not descs:
        flash("At least one item is required.", "danger")
        return redirect(url_for('documents'))
    doc = Document(
        doc_type        = 'PI',
        quote_no        = next_quote_number(),
        doc_date        = _parse_date(request.form.get('doc_date')) or date.today(),
        dispatch_from   = disp,
        customer_id     = int(cid),
        gst_applied     = request.form.get('gst_applied') == 'on',
        freight_charges = float(request.form.get('freight_charges') or 0),
        follow_up_date  = _parse_date(request.form.get('follow_up_date')) or (date.today()+timedelta(days=7)),
        status          = request.form.get('status','pending'),
        lost_reason     = request.form.get('lost_reason','').strip(),
        notes           = request.form.get('notes','').strip(),
        created_by_id   = current_user.id,
    )
    db.session.add(doc); db.session.flush()
    _save_line_items(doc, request.form)
    doc.recalc()
    db.session.commit()
    flash(f"PI {doc.quote_no} created with {len(doc.line_items)} item(s).", "success")
    return redirect(url_for('documents'))


@app.route('/pi/<int:doc_id>/edit', methods=['GET','POST'])
@login_required
def edit_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("Access denied.", "danger"); return redirect(url_for('documents'))
    custs = visible_custs().order_by(Customer.name).all()
    items = Item.query.order_by(Item.name).all()
    items_data = {it.name: {'uom': it.default_uom or '', 'price': it.default_price or 0} for it in items}
    if request.method == 'POST':
        doc.doc_type        = 'PI'
        doc.doc_date        = _parse_date(request.form.get('doc_date')) or doc.doc_date
        doc.dispatch_from   = request.form.get('dispatch_from', doc.dispatch_from)
        doc.customer_id     = int(request.form.get('customer_id', doc.customer_id))
        doc.gst_applied     = request.form.get('gst_applied') == 'on'
        doc.freight_charges = float(request.form.get('freight_charges') or 0)
        doc.follow_up_date  = _parse_date(request.form.get('follow_up_date'))
        doc.status          = request.form.get('status', doc.status)
        doc.lost_reason     = request.form.get('lost_reason','').strip()
        doc.notes           = request.form.get('notes','').strip()
        _save_line_items(doc, request.form)
        doc.recalc()
        db.session.commit()
        flash(f"PI {doc.quote_no} updated.", "success")
        return redirect(url_for('documents'))
    return render_template('edit_document.html', doc=doc,
        customers=custs, items=items, items_data=items_data,
        dispatch_units=DISPATCH_UNITS, uom_options=UOM_OPTIONS,
        doc_statuses=DOC_STATUSES, status_labels=STATUS_LABELS)


@app.route('/pi/<int:doc_id>/delete', methods=['POST'])
@login_required
def delete_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("Access denied.", "danger"); return redirect(url_for('documents'))
    db.session.delete(doc); db.session.commit()
    flash("PI deleted.", "success"); return redirect(url_for('documents'))


@app.route('/pi/<int:doc_id>/pdf')
@login_required
def download_document_pdf(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("Access denied.", "danger"); return redirect(url_for('documents'))
    buf      = generate_pi_pdf(doc)
    filename = f"PI_{doc.quote_no}_{doc.customer.name[:20].replace(' ','_')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@app.route('/pi/<int:doc_id>/duplicate', methods=['POST'])
@login_required
def duplicate_document(doc_id):
    src = Document.query.get_or_404(doc_id)
    new_doc = Document(
        doc_type=src.doc_type, quote_no=next_quote_number(),
        doc_date=date.today(), dispatch_from=src.dispatch_from,
        customer_id=src.customer_id, gst_applied=src.gst_applied,
        freight_charges=src.freight_charges, notes=src.notes,
        status='pending',
        follow_up_date=date.today()+timedelta(days=7),
        created_by_id=current_user.id)
    db.session.add(new_doc); db.session.flush()
    for li in src.line_items:
        db.session.add(DocumentItem(document_id=new_doc.id, item_desc=li.item_desc,
            packaging=li.packaging, qty=li.qty, uom=li.uom, price=li.price))
    new_doc.recalc(); db.session.commit()
    flash(f"Duplicated as {new_doc.quote_no}.", "success")
    return redirect(url_for('documents'))


@app.route('/pi/export')
@login_required
def export_documents():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    docs = visible_docs().order_by(Document.created_at.desc()).all()
    wb = Workbook(); ws = wb.active; ws.title = "PI Export"
    headers = ['Quote No.','Date','Dispatch','Customer','Customer Code','Item','Pkg','Qty','Total Qty',
               'UOM','Price','Line Total','Base Amt','GST','Freight','Total','Status','Lost Reason',
               'Follow-up','Notes','Created By']
    hf = PatternFill("solid", fgColor="0B2545"); hfont = Font(bold=True, color="FFFFFF", size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
    ri = 2
    for d in docs:
        items = d.line_items or []
        if not items:
            items = [type('L', (), {'item_desc': d.item_desc,'packaging':d.packaging,'qty':d.qty,
                                    'uom':d.uom,'price':d.price,'total_qty':0,'line_total':0})()]
        for idx, li in enumerate(items):
            row = [
                d.quote_no if idx==0 else '',
                d.doc_date.strftime('%d-%b-%Y') if idx==0 else '',
                d.dispatch_from or '' if idx==0 else '',
                d.customer.name if (idx==0 and d.customer) else '',
                d.customer.code if (idx==0 and d.customer) else '',
                li.item_desc,
                li.packaging,
                li.qty,
                getattr(li,'total_qty', (li.packaging or 0)*(li.qty or 0)),
                li.uom,
                li.price,
                getattr(li,'line_total', (li.qty or 0)*(li.price or 0)),
                d.base_amount if idx==0 else '',
                d.gst_amount  if idx==0 else '',
                d.freight_charges if idx==0 else '',
                d.total_amount if idx==0 else '',
                d.status_label if idx==0 else '',
                d.lost_reason or '' if idx==0 else '',
                d.follow_up_date.strftime('%d-%b-%Y') if (idx==0 and d.follow_up_date) else '',
                d.notes or '' if idx==0 else '',
                d.created_by.full_name if (idx==0 and d.created_by) else '',
            ]
            for ci, val in enumerate(row, 1): ws.cell(row=ri, column=ci, value=val)
            if d.status=='delivered':
                for ci in range(1,len(headers)+1): ws.cell(row=ri,column=ci).fill=PatternFill("solid",fgColor="DCFCE7")
            elif d.status=='lost':
                for ci in range(1,len(headers)+1): ws.cell(row=ri,column=ci).fill=PatternFill("solid",fgColor="FCE7F3")
            ri += 1
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 15
    ws.column_dimensions['D'].width = 28; ws.column_dimensions['F'].width = 32
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"PI_Export_{date.today().strftime('%d%b%Y')}.xlsx")

# ── CUSTOMERS ─────────────────────────────────────────────────────────────────

@app.route('/customers')
@login_required
def customers():
    return render_template('customers.html', customers=visible_custs().order_by(Customer.name).all())


@app.route('/customers/new', methods=['POST'])
@login_required
def new_customer():
    name = request.form.get('name','').strip()
    c = Customer(
        name=name or 'Unnamed',
        code=next_customer_code(),
        contact_person=request.form.get('contact_person','').strip(),
        phone=request.form.get('phone','').strip(),
        gst_number=request.form.get('gst_number','').strip(),
        created_by_id=current_user.id)
    db.session.add(c); db.session.commit()
    flash(f"Customer '{c.name}' added (Code: {c.code}).", "success")
    return redirect(url_for('customers'))


@app.route('/customers/<int:cust_id>/delete', methods=['POST'])
@login_required
def delete_customer(cust_id):
    c = Customer.query.get_or_404(cust_id)
    if not current_user.is_admin() and c.created_by_id != current_user.id:
        flash("Access denied.", "danger"); return redirect(url_for('customers'))
    db.session.delete(c); db.session.commit()
    flash("Customer deleted.", "success"); return redirect(url_for('customers'))


@app.route('/customers/bulk-import', methods=['POST'])
@login_required
def bulk_import_customers():
    file = request.files.get('file')
    if not file or not file.filename:
        flash("Choose a file.", "danger"); return redirect(url_for('customers'))
    rows = _read_tabular_file(file)
    if rows is None:
        flash("Use .csv or .xlsx", "danger"); return redirect(url_for('customers'))
    n = 0
    for row in rows:
        name = (row.get('name') or row.get('Name') or row.get('Customer Name') or '').strip()
        if not name: continue
        db.session.add(Customer(
            name=name, code=next_customer_code(),
            contact_person=(row.get('contact_person') or row.get('Contact Person') or '').strip(),
            phone=(row.get('phone') or row.get('Phone') or '').strip(),
            gst_number=(row.get('gst_number') or row.get('GST') or '').strip(),
            created_by_id=current_user.id)); n += 1
    db.session.commit()
    flash(f"Imported {n} customers.", "success"); return redirect(url_for('customers'))


@app.route('/customers/template')
@login_required
def customer_template():
    return _csv_response(['name','contact_person','phone','gst_number'],
        [['Example Traders Pvt Ltd','Ramesh Kumar','+919876543210','29ABCDE1234F1Z5']],
        'customer_template.csv')


@app.route('/customers/<int:cust_id>/statement')
@login_required
def customer_statement(cust_id):
    c = Customer.query.get_or_404(cust_id)
    if not current_user.is_admin() and c.created_by_id != current_user.id:
        flash("Access denied.","danger"); return redirect(url_for('customers'))
    docs  = Document.query.filter_by(customer_id=cust_id).order_by(Document.doc_date.desc()).all()
    total = sum(d.total_amount or 0 for d in docs)
    deliv = sum(d.total_amount or 0 for d in docs if d.status=='delivered')
    pend  = sum(d.total_amount or 0 for d in docs if d.status not in ('delivered','lost'))
    return render_template('customer_statement.html',
        customer=c, docs=docs, total=total, delivered=deliv,
        pending=pend, lost_count=len([d for d in docs if d.status=='lost']))

# ── ITEMS ─────────────────────────────────────────────────────────────────────

@app.route('/items')
@login_required
def items():
    all_items = Item.query.order_by(Item.name).all()
    return render_template('items.html', items=all_items, uom_options=UOM_OPTIONS)


@app.route('/items/new', methods=['POST'])
@login_required
def new_item():
    name = request.form.get('name','').strip()
    if not name:
        flash("Name required.","danger"); return redirect(url_for('items'))
    db.session.add(Item(name=name, default_uom=request.form.get('default_uom'),
        default_price=float(request.form.get('default_price') or 0),
        created_by_id=current_user.id))
    db.session.commit(); flash(f"Item '{name}' added.","success")
    return redirect(url_for('items'))


@app.route('/items/new-multi', methods=['POST'])
@login_required
def new_item_multi():
    names  = request.form.getlist('name[]')
    uoms   = request.form.getlist('uom[]')
    prices = request.form.getlist('price[]')
    n = 0
    for i, name in enumerate(names):
        name = name.strip()
        if not name: continue
        db.session.add(Item(name=name,
            default_uom=uoms[i] if i<len(uoms) else '',
            default_price=float(prices[i] if i<len(prices) and prices[i] else 0),
            created_by_id=current_user.id)); n += 1
    db.session.commit(); flash(f"{n} item(s) added.","success")
    return redirect(url_for('items'))


@app.route('/items/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_item(item_id):
    db.session.delete(Item.query.get_or_404(item_id))
    db.session.commit(); flash("Item deleted.","success")
    return redirect(url_for('items'))


@app.route('/items/bulk-import', methods=['POST'])
@login_required
def bulk_import_items():
    file = request.files.get('file')
    if not file or not file.filename:
        flash("Choose a file.","danger"); return redirect(url_for('items'))
    rows = _read_tabular_file(file)
    if rows is None:
        flash("Use .csv or .xlsx","danger"); return redirect(url_for('items'))
    n = 0
    for row in rows:
        name = (row.get('name') or row.get('Name') or '').strip()
        if not name: continue
        db.session.add(Item(name=name,
            default_uom=(row.get('uom') or row.get('UOM') or '').strip(),
            default_price=float(row.get('price') or row.get('Price') or 0),
            created_by_id=current_user.id)); n += 1
    db.session.commit(); flash(f"Imported {n} items.","success")
    return redirect(url_for('items'))


@app.route('/items/template')
@login_required
def item_template():
    return _csv_response(['name','uom','price'],[['28mm Alaska cap','Nos','0.32']],'item_template.csv')


@app.route('/api/item-price')
@login_required
def api_item_price():
    name = request.args.get('name','')
    it   = Item.query.filter_by(name=name).first()
    if it: return jsonify({'uom': it.default_uom or '', 'price': it.default_price or 0})
    return jsonify({'uom':'','price':0})

# ── FOLLOW-UPS ────────────────────────────────────────────────────────────────

@app.route('/followups')
@login_required
def followups():
    today    = date.today()
    all_docs = visible_docs().filter(
        Document.status.notin_(['delivered','lost']),
        Document.follow_up_date != None
    ).order_by(Document.follow_up_date).all()
    overdue   = [d for d in all_docs if d.follow_up_date < today]
    due_today = [d for d in all_docs if d.follow_up_date == today]
    upcoming  = [d for d in all_docs if d.follow_up_date > today]
    return render_template('followups.html', overdue=overdue, due_today=due_today, upcoming=upcoming, today=today)

# ── SEARCH ────────────────────────────────────────────────────────────────────

@app.route('/search')
@login_required
def search():
    q = request.args.get('q','').strip()
    if not q:
        return render_template('search.html', q='', doc_results=[], cust_results=[], item_results=[])
    ql = q.lower()
    doc_results  = [d for d in visible_docs().all() if
                    ql in (d.customer.name.lower() if d.customer else '')
                    or ql in d.quote_no.lower()
                    or any(ql in (li.item_desc or '').lower() for li in d.line_items)
                    or ql in (d.item_desc or '').lower()]
    cust_results = [c for c in visible_custs().all() if
                    ql in c.name.lower() or ql in (c.code or '').lower()
                    or ql in (c.contact_person or '').lower() or ql in (c.phone or '').lower()]
    item_results = [i for i in Item.query.all() if ql in i.name.lower()]
    return render_template('search.html', q=q,
        doc_results=doc_results, cust_results=cust_results, item_results=item_results)

# ── ADMIN: USERS ──────────────────────────────────────────────────────────────

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    return render_template('admin_users.html', users=User.query.order_by(User.username).all())


@app.route('/admin/users/new', methods=['POST'])
@login_required
@admin_required
def admin_new_user():
    username  = request.form.get('username','').strip().lower()
    full_name = request.form.get('full_name','').strip()
    password  = request.form.get('password','')
    role      = request.form.get('role','staff')
    if not username or not password or not full_name:
        flash("Username, full name and password required.","danger"); return redirect(url_for('admin_users'))
    if User.query.filter_by(username=username).first():
        flash("Username already exists.","danger"); return redirect(url_for('admin_users'))
    u = User(username=username, full_name=full_name, role=role)
    u.set_password(password); db.session.add(u); db.session.commit()
    flash(f"User '{username}' created.","success"); return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user(user_id):
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash("Cannot deactivate own account.","danger"); return redirect(url_for('admin_users'))
    u.active = not u.active; db.session.commit()
    flash(f"User '{u.username}' {'activated' if u.active else 'deactivated'}.","success")
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    u = User.query.get_or_404(user_id)
    pw = request.form.get('new_password','')
    if len(pw) < 6:
        flash("Min 6 characters.","danger"); return redirect(url_for('admin_users'))
    u.set_password(pw); db.session.commit()
    flash(f"Password reset for '{u.username}'.","success"); return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id or u.role == 'admin':
        flash("Cannot delete this account.","danger"); return redirect(url_for('admin_users'))
    db.session.delete(u); db.session.commit()
    flash(f"User '{u.username}' deleted.","success"); return redirect(url_for('admin_users'))

# ── ADMIN: REPORTS ────────────────────────────────────────────────────────────

@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    today = date.today()
    sel_month = request.args.get('month', type=int, default=today.month)
    sel_year  = request.args.get('year',  type=int, default=today.year)
    sel_uid   = request.args.get('user_id', type=int)

    staff_users = User.query.filter_by(role='staff').order_by(User.full_name).all()
    months_list = [(y, m) for y in range(today.year, today.year-2, -1) for m in range(12,0,-1)]

    def month_docs(uid):
        return Document.query.filter_by(created_by_id=uid).filter(
            db.extract('year',  Document.doc_date) == sel_year,
            db.extract('month', Document.doc_date) == sel_month,
        ).all()

    def get_target(uid):
        t = UserMonthlyTarget.query.filter_by(user_id=uid, year=sel_year, month=sel_month).first()
        return t.target if t else 0

    summaries = []
    for u in staff_users:
        udocs   = month_docs(u.id)
        deliv   = sum(d.total_amount or 0 for d in udocs if d.status=='delivered')
        target  = get_target(u.id)
        summaries.append(dict(
            user=u,
            total_pi      = len(udocs),
            total_converted = len([d for d in udocs if d.status=='delivered']),
            delivered_val = deliv,
            pending_val   = sum(d.total_amount or 0 for d in udocs if d.status not in ('delivered','lost')),
            lost_count    = len([d for d in udocs if d.status=='lost']),
            target        = target,
            target_pct    = round(deliv/target*100,1) if target else 0,
        ))

    detail_docs = []
    sel_user    = None
    if sel_uid:
        sel_user    = User.query.get(sel_uid)
        detail_docs = month_docs(sel_uid)

    return render_template('admin_reports.html',
        summaries=summaries, staff_users=staff_users,
        sel_month=sel_month, sel_year=sel_year,
        sel_user=sel_user, detail_docs=detail_docs,
        months_list=months_list,
        month_name=calendar.month_name[sel_month],
        today=today)


@app.route('/admin/reports/set-target', methods=['POST'])
@login_required
@admin_required
def admin_set_target():
    uid    = int(request.form.get('user_id'))
    month  = int(request.form.get('month'))
    year   = int(request.form.get('year'))
    target = float(request.form.get('target') or 0)
    t = UserMonthlyTarget.query.filter_by(user_id=uid, year=year, month=month).first()
    if t:
        t.target = target
    else:
        db.session.add(UserMonthlyTarget(user_id=uid, year=year, month=month, target=target))
    db.session.commit()
    flash("Target updated.","success")
    return redirect(url_for('admin_reports', month=month, year=year))


@app.route('/admin/reports/export')
@login_required
@admin_required
def export_staff_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    today     = date.today()
    sel_month = request.args.get('month', type=int, default=today.month)
    sel_year  = request.args.get('year',  type=int, default=today.year)
    sel_uid   = request.args.get('user_id', type=int)
    month_name = calendar.month_name[sel_month]

    wb = Workbook(); ws = wb.active
    ws.title = f"{month_name} {sel_year}"
    hf = PatternFill("solid", fgColor="0B2545")
    hfont = Font(bold=True, color="FFFFFF", size=10)

    if sel_uid:
        # Individual staff report
        u = User.query.get_or_404(sel_uid)
        docs = Document.query.filter_by(created_by_id=sel_uid).filter(
            db.extract('year',Document.doc_date)==sel_year,
            db.extract('month',Document.doc_date)==sel_month).all()
        ws.append([f"Staff Report: {u.full_name} — {month_name} {sel_year}"])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])
        headers = ['Quote No.','Date','Customer','Items','Total','Status','Follow-up','Lost Reason']
        for ci, h in enumerate(headers,1):
            cell = ws.cell(row=3,column=ci,value=h); cell.fill=hf; cell.font=hfont
        for ri, d in enumerate(docs, 4):
            ws.append([d.quote_no, d.doc_date.strftime('%d-%b-%Y') if d.doc_date else '',
                       d.customer.name if d.customer else '', d.items_summary,
                       d.total_amount, d.status_label,
                       d.follow_up_date.strftime('%d-%b-%Y') if d.follow_up_date else '',
                       d.lost_reason or ''])
    else:
        # Full staff summary
        ws.append([f"Staff Summary Report — {month_name} {sel_year}"])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])
        headers = ['Staff Name','Total PI','Converted','Delivered Value','Pending Value','Lost','Target','Achieved %']
        for ci, h in enumerate(headers,1):
            cell = ws.cell(row=3,column=ci,value=h); cell.fill=hf; cell.font=hfont
        for u in User.query.filter_by(role='staff').order_by(User.full_name).all():
            docs  = Document.query.filter_by(created_by_id=u.id).filter(
                db.extract('year',Document.doc_date)==sel_year,
                db.extract('month',Document.doc_date)==sel_month).all()
            deliv = sum(d.total_amount or 0 for d in docs if d.status=='delivered')
            tgt   = UserMonthlyTarget.query.filter_by(user_id=u.id,year=sel_year,month=sel_month).first()
            target= tgt.target if tgt else 0
            pct   = round(deliv/target*100,1) if target else 0
            ws.append([u.full_name, len(docs),
                       len([d for d in docs if d.status=='delivered']),
                       deliv,
                       sum(d.total_amount or 0 for d in docs if d.status not in ('delivered','lost')),
                       len([d for d in docs if d.status=='lost']),
                       target, pct])
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    suffix = f"_{User.query.get(sel_uid).username}" if sel_uid else "_all"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f"Report_{month_name}{sel_year}{suffix}.xlsx")

# ── HELPERS ───────────────────────────────────────────────────────────────────

def _read_tabular_file(fs):
    fn = fs.filename.lower()
    if fn.endswith('.csv'):
        return list(csv.DictReader(io.StringIO(fs.stream.read().decode('utf-8-sig'))))
    if fn.endswith('.xlsx'):
        wb = load_workbook(fs, read_only=True, data_only=True); ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try: hdr = [str(h).strip() if h else '' for h in next(rows)]
        except StopIteration: return []
        return [{hdr[i]:('' if i>=len(r) or r[i] is None else str(r[i])) for i in range(len(hdr))} for r in rows]
    return None


def _csv_response(headers, rows, filename):
    out = io.StringIO(); csv.writer(out).writerow(headers); csv.writer(out).writerows(rows)
    return Response(out.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'})

# ── STARTUP ───────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    is_pg = 'postgresql' in str(db.engine.url)
    with db.engine.connect() as conn:
        stmts = ([
            'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS monthly_target FLOAT DEFAULT 0',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS lost_reason TEXT',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS item_desc VARCHAR(300)',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS packaging VARCHAR(50)',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS qty FLOAT DEFAULT 0',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS uom VARCHAR(20)',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS price FLOAT DEFAULT 0',
            'ALTER TABLE document ALTER COLUMN item_desc DROP NOT NULL',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS code VARCHAR(20)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS gst_number VARCHAR(20)',
            'ALTER TABLE document_item ALTER COLUMN packaging TYPE FLOAT USING packaging::FLOAT',
        ] if is_pg else [
            'ALTER TABLE "user" ADD COLUMN monthly_target FLOAT DEFAULT 0',
            'ALTER TABLE document ADD COLUMN lost_reason TEXT',
            'ALTER TABLE document ADD COLUMN item_desc VARCHAR(300)',
            'ALTER TABLE document ADD COLUMN packaging VARCHAR(50)',
            'ALTER TABLE document ADD COLUMN qty FLOAT DEFAULT 0',
            'ALTER TABLE document ADD COLUMN uom VARCHAR(20)',
            'ALTER TABLE document ADD COLUMN price FLOAT DEFAULT 0',
            'ALTER TABLE customer ADD COLUMN code VARCHAR(20)',
            'ALTER TABLE customer ADD COLUMN gst_number VARCHAR(20)',
        ])
        for stmt in stmts:
            try: conn.execute(db.text(stmt)); conn.commit()
            except: conn.rollback()
    if not User.query.filter_by(username='admin').first():
        a = User(username='admin', full_name='Administrator', role='admin')
        a.set_password('admin123'); db.session.add(a); db.session.commit()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT',5000)), debug=False)
